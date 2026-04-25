"""Excel report export using openpyxl."""
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


BRAND_HEX    = "1877F2"
HEADER_HEX   = "1877F2"
ALT_ROW_HEX  = "F0F2F5"
GREEN_HEX    = "2E7D32"
RED_HEX      = "C62828"
ORANGE_HEX   = "E65C00"


def _header_style(ws, row_num, cols, color_hex=HEADER_HEX):
    fill = PatternFill("solid", fgColor=color_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align


def _alt_row_style(ws, row_num, cols, even: bool):
    if even:
        fill = PatternFill("solid", fgColor=ALT_ROW_HEX)
        for c in range(1, cols + 1):
            ws.cell(row=row_num, column=c).fill = fill


def _kpi_cell(ws, row, col, label, value, color_hex=BRAND_HEX):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(bold=True, color="888888", size=8)
    lc.alignment = Alignment(horizontal="center")

    vc = ws.cell(row=row + 1, column=col, value=value)
    vc.font = Font(bold=True, color=color_hex, size=14)
    vc.alignment = Alignment(horizontal="center")

    for r in [row, row + 1]:
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FFFFFF")


def generate_excel(totals: dict, perf_df, time_df,
                   alerts_df=None, brand_name="Meta Ads") -> bytes:
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed.")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{brand_name} – Performance Report"
    title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor=BRAND_HEX)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    sub.font  = Font(italic=True, size=9, color="606770")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # KPI row
    kpis = [
        ("TOTAL SPEND",    f"Rs {totals.get('total_spend',0):,.0f}", BRAND_HEX),
        ("TOTAL CLICKS",   f"{totals.get('total_clicks',0):,}",      "E65C00"),
        ("TOTAL LEADS",    f"{totals.get('total_leads',0):,}",        GREEN_HEX),
        ("AVG CTR",        f"{totals.get('avg_ctr',0):.2f}%",        "9C27B0"),
        ("TOTAL IMPR.",    f"{totals.get('total_impressions',0):,}",  "00BCD4"),
    ]
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 24
    for i, (lbl, val, col) in enumerate(kpis, 1):
        _kpi_cell(ws, 4, i, lbl, val, col)

    # ── Sheet 2: Campaigns ────────────────────────────────────
    wc = wb.create_sheet("Campaigns")
    wc.sheet_view.showGridLines = False
    headers = ["Campaign Name", "Status", "Objective", "Total Spend",
               "Avg ROAS", "Avg CTR", "Clicks", "Leads", "Impressions"]
    for c, h in enumerate(headers, 1):
        wc.cell(row=1, column=c, value=h)
    _header_style(wc, 1, len(headers))
    wc.row_dimensions[1].height = 20

    if not perf_df.empty:
        for i, (_, r) in enumerate(perf_df.iterrows(), 2):
            row_data = [
                r.get("name",""),
                r.get("status",""),
                str(r.get("objective","")).replace("OUTCOME_",""),
                round(float(r.get("total_spend",0)),2),
                round(float(r.get("avg_roas",0)),2),
                round(float(r.get("avg_ctr",0)),2),
                int(r.get("clicks",0)),
                int(r.get("leads",0)),
                int(r.get("impressions",0)),
            ]
            for c, v in enumerate(row_data, 1):
                wc.cell(row=i, column=c, value=v)
            _alt_row_style(wc, i, len(headers), i % 2 == 0)
            # Color status
            status_cell = wc.cell(row=i, column=2)
            if r.get("status") == "ACTIVE":
                status_cell.font = Font(color=GREEN_HEX, bold=True)
            elif r.get("status") == "PAUSED":
                status_cell.font = Font(color=ORANGE_HEX, bold=True)

    for col in wc.columns:
        wc.column_dimensions[get_column_letter(col[0].column)].width = 18

    # ── Sheet 3: Daily Data ───────────────────────────────────
    wd = wb.create_sheet("Daily Data")
    wd.sheet_view.showGridLines = False
    d_headers = ["Date","Spend","Clicks","Impressions","ROAS","CTR","Leads"]
    for c, h in enumerate(d_headers, 1):
        wd.cell(row=1, column=c, value=h)
    _header_style(wd, 1, len(d_headers))

    if not time_df.empty:
        for i, (_, r) in enumerate(time_df.iterrows(), 2):
            row_data = [
                str(r.get("date",""))[:10],
                round(float(r.get("spend",0)),2),
                int(r.get("clicks",0)),
                int(r.get("impressions",0)),
                round(float(r.get("roas",0)),2),
                round(float(r.get("ctr",0)),2),
                int(r.get("leads",0)) if "leads" in r else 0,
            ]
            for c, v in enumerate(row_data, 1):
                wd.cell(row=i, column=c, value=v)
            _alt_row_style(wd, i, len(d_headers), i % 2 == 0)
    for col in wd.columns:
        wd.column_dimensions[get_column_letter(col[0].column)].width = 15

    # ── Sheet 4: Alerts ───────────────────────────────────────
    if alerts_df is not None and not alerts_df.empty:
        wa = wb.create_sheet("Alerts")
        wa.sheet_view.showGridLines = False
        a_headers = ["Created At","Alert Type","Message","Campaign ID"]
        for c, h in enumerate(a_headers, 1):
            wa.cell(row=1, column=c, value=h)
        _header_style(wa, 1, len(a_headers), RED_HEX)
        for i, (_, r) in enumerate(alerts_df.iterrows(), 2):
            for c, key in enumerate(["created_at","alert_type","message","campaign_id"], 1):
                wa.cell(row=i, column=c, value=str(r.get(key,"")))
            _alt_row_style(wa, i, len(a_headers), i % 2 == 0)
        for col in wa.columns:
            wa.column_dimensions[get_column_letter(col[0].column)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
